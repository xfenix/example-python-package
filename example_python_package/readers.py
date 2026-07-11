"""Reading seam: a Protocol plus CSV and XLSX implementations.

Two inference/write passes require re-openable readers, so each reader holds a
path and yields a fresh iterator on every ``iter_rows`` call.
"""

import csv
import dataclasses
import typing

import openpyxl

import example_python_package.errors as errors_module


if typing.TYPE_CHECKING:
    import collections.abc
    import pathlib


COLUMN_NAME_PREFIX: typing.Final = "col_"
CSV_SUFFIX: typing.Final = ".csv"
XLSX_SUFFIX: typing.Final = ".xlsx"


@typing.runtime_checkable
class RowReader(typing.Protocol):
    @property
    def header_names(self) -> tuple[str, ...]: ...

    def iter_rows(self) -> collections.abc.Iterator[tuple[object, ...]]: ...


@typing.final
@dataclasses.dataclass(kw_only=True, slots=True, frozen=True)
class CsvReader:
    file_path: pathlib.Path
    header_names: tuple[str, ...]
    has_header: bool

    def iter_rows(self) -> collections.abc.Iterator[tuple[object, ...]]:
        with self.file_path.open(newline="", encoding="utf-8-sig") as csv_handle:
            row_iterator: typing.Final = csv.reader(csv_handle)
            if self.has_header:
                next(row_iterator, None)
            for one_parsed_row in row_iterator:
                yield tuple(one_parsed_row)


@typing.final
@dataclasses.dataclass(kw_only=True, slots=True, frozen=True)
class XlsxReader:
    file_path: pathlib.Path
    header_names: tuple[str, ...]
    has_header: bool

    def iter_rows(self) -> collections.abc.Iterator[tuple[object, ...]]:
        workbook: typing.Final = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        try:
            row_iterator: typing.Final = workbook.active.iter_rows(values_only=True)
            if self.has_header:
                next(row_iterator, None)
            yield from row_iterator
        finally:
            workbook.close()


def build_reader(file_path: pathlib.Path, *, has_header: bool = True) -> RowReader:
    file_suffix: typing.Final = file_path.suffix.lower()
    if file_suffix == CSV_SUFFIX:
        return build_csv_reader(file_path, has_header=has_header)
    if file_suffix == XLSX_SUFFIX:
        return build_xlsx_reader(file_path, has_header=has_header)
    raise errors_module.UnsupportedFormatError(f"No reader for extension {file_path.suffix!r}.")


def build_csv_reader(file_path: pathlib.Path, *, has_header: bool) -> CsvReader:
    return CsvReader(
        file_path=file_path,
        header_names=resolve_header_names(read_first_csv_row(file_path), has_header=has_header),
        has_header=has_header,
    )


def build_xlsx_reader(file_path: pathlib.Path, *, has_header: bool) -> XlsxReader:
    return XlsxReader(
        file_path=file_path,
        header_names=resolve_header_names(read_first_xlsx_row(file_path), has_header=has_header),
        has_header=has_header,
    )


def read_first_csv_row(file_path: pathlib.Path) -> tuple[object, ...]:
    with file_path.open(newline="", encoding="utf-8-sig") as csv_handle:
        return tuple(next(csv.reader(csv_handle), []))


def read_first_xlsx_row(file_path: pathlib.Path) -> tuple[object, ...]:
    workbook: typing.Final = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        return tuple(next(workbook.active.iter_rows(values_only=True), ()))
    finally:
        workbook.close()


def resolve_header_names(first_row: tuple[object, ...], *, has_header: bool) -> tuple[str, ...]:
    if not has_header:
        return make_synthetic_names(len(first_row))
    return make_unique_names(
        tuple(
            make_column_name(one_cell_value, one_position)
            for one_position, one_cell_value in enumerate(first_row, start=1)
        ),
    )


def make_synthetic_names(column_count: int) -> tuple[str, ...]:
    return tuple(f"{COLUMN_NAME_PREFIX}{one_position}" for one_position in range(1, column_count + 1))


def make_column_name(cell_value: object, position: int) -> str:
    if cell_value is None:
        return f"{COLUMN_NAME_PREFIX}{position}"
    return str(cell_value).strip() or f"{COLUMN_NAME_PREFIX}{position}"


def make_unique_names(raw_names: tuple[str, ...]) -> tuple[str, ...]:
    seen_counts: typing.Final[dict[str, int]] = {}
    unique_names: typing.Final[list[str]] = []
    for one_candidate_name in raw_names:
        seen_counts[one_candidate_name] = seen_counts.get(one_candidate_name, 0) + 1
        occurrence = seen_counts[one_candidate_name]
        unique_names.append(one_candidate_name if occurrence == 1 else f"{one_candidate_name}_{occurrence}")
    return tuple(unique_names)
